import requests
from lxml import etree
from openpyxl import Workbook
import sys
import io


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36'}

def get_first_text(list):
    try:
        return list[0].strip()
    except:
        return ""

urls = ['https://www.8world.com/southeast-asia?page={}'.format(str(i * 1)) for i in range(100)]
k= "https://www.8world.com"

keywords = ['洪水']  # 设置关键词列表

workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'flood news of Malaysia'
worksheet['A1'] = 'article'
worksheet['B1'] = 'time'
worksheet['C1'] = 'link'
count = 1

for url in urls:
    res = requests.get(url=url, headers=headers)
    html = etree.HTML(res.text)
    lis = html.xpath(
        '/html/body/div[1]/div[2]/main/div[1]/div/article/div/div[2]/div/div/div[1]/div/div/div/div/div/div/div/div/div/div/article')


    for li in lis:
        title = get_first_text(li.xpath('./div/div/div[1]/h3/a/span/text()'))
        time = get_first_text(li.xpath('./div/div/div[3]/time/text()'))
        link = get_first_text(li.xpath('./div/div/div[1]/h3/a/@href'))

        if any(keyword in title.lower() for keyword in keywords):
           worksheet.cell(row=count + 1, column=1, value=title)
           worksheet.cell(row=count + 1, column=2, value=time)
           worksheet.cell(row=count + 1, column=3, value=k + link)
           count += 1



workbook.save('news.xlsx')

